import re

from app.services import api_client

ADMIN_TOKENS = {"access_token": "a", "refresh_token": "r", "token_type": "bearer"}
ADMIN_ME = {
    "id_usuario": 1, "nombre": "Admin", "apellido_paterno": "Sistema",
    "apellido_materno": None, "correo": "admin@cafeteria.com",
    "nombre_usuario": "admin", "id_rol": 1, "activo": True,
    "fecha_registro": "2026-07-04T00:00:00Z",
    "rol": {"id_rol": 1, "nombre_rol": "Administrador", "descripcion": None},
}
VENTAS = [{"folio": "V-0001", "fecha": "2026-07-05T12:00:00Z", "mesa": 3,
           "total": "232.00", "metodos": "Efectivo"}]
GASTOS = [{"fecha": "2026-07-05T09:00:00Z", "categoria": "Servicios",
           "concepto": "Luz", "monto": "250.00"}]
INVENTARIO = [{"nombre": "Café", "unidad": "kg", "stock_actual": "5.00",
               "stock_minimo": "10.00", "nivel_pct": 50, "bajo_minimo": True}]
ESTADO_RESULTADOS = [
    {"periodo": "2026-07", "ventas": "1000.00", "gastos": "400.00",
     "compras": "200.00", "utilidad": "600.00"},
]
USUARIOS = [
    {"id_usuario": 7, "nombre": "Juan", "apellido_paterno": "Pérez",
     "apellido_materno": None},
]
METODOS = [{"id_metodo_pago": 2, "nombre_metodo": "Tarjeta"}]
CATEGORIAS = [{"id_categoria_gasto": 3, "nombre_categoria": "Servicios"}]


def _login(client, monkeypatch):
    monkeypatch.setattr(api_client, "login", lambda c, p: ADMIN_TOKENS)
    monkeypatch.setattr(api_client, "get_me", lambda a: ADMIN_ME)
    client.post("/login", data={"correo": "admin@cafeteria.com", "password": "x"})


def _stub(monkeypatch, capturados=None):
    """Stub de los 4 endpoints de reporte + dropdowns.

    Los stubs de reporte capturan los filtros recibidos en `capturados`
    (dict mutable) para que los tests verifiquen que la ruta los propaga
    tal cual al api_client. Firmas explícitas (no *args genérico) para que
    un cambio de orden de parámetros rompa el test, no lo esconda.
    """
    if capturados is None:
        capturados = {}

    def _ventas(a, desde=None, hasta=None, id_usuario=None, id_metodo=None):
        capturados["ventas"] = {"id_usuario": id_usuario, "id_metodo": id_metodo}
        return VENTAS

    def _gastos(a, desde=None, hasta=None, id_usuario=None, id_categoria=None):
        capturados["gastos"] = {"id_usuario": id_usuario, "id_categoria": id_categoria}
        return GASTOS

    def _inventario(a, solo_bajo_minimo=False):
        capturados["inventario"] = {"solo_bajo_minimo": solo_bajo_minimo}
        return INVENTARIO

    def _estado_resultados(a, desde=None, hasta=None, agrupar="dia"):
        capturados["estado_resultados"] = {"agrupar": agrupar}
        return ESTADO_RESULTADOS

    monkeypatch.setattr(api_client, "get_reporte_ventas", _ventas)
    monkeypatch.setattr(api_client, "get_reporte_gastos", _gastos)
    monkeypatch.setattr(api_client, "get_inventario_niveles", _inventario)
    monkeypatch.setattr(api_client, "get_estado_resultados", _estado_resultados)
    monkeypatch.setattr(api_client, "list_usuarios", lambda a, q=None: USUARIOS)
    monkeypatch.setattr(api_client, "get_metodos_pago", lambda a: METODOS)
    monkeypatch.setattr(api_client, "get_gastos_categorias", lambda a: CATEGORIAS)
    return capturados


def test_reportes_sin_sesion_redirige(client):
    r = client.get("/reportes")
    assert r.status_code == 302
    assert "/login" in r.headers["Location"]


def test_reportes_preview_ventas(client, monkeypatch):
    _login(client, monkeypatch)
    _stub(monkeypatch)
    cuerpo = client.get("/reportes?tipo=ventas").get_data(as_text=True)
    assert "V-0001" in cuerpo
    assert "Efectivo" in cuerpo
    assert "Reporte de Ventas" in cuerpo


def test_reportes_preview_gastos(client, monkeypatch):
    _login(client, monkeypatch)
    _stub(monkeypatch)
    cuerpo = client.get("/reportes?tipo=gastos").get_data(as_text=True)
    assert "Luz" in cuerpo
    assert "Reporte de Gastos" in cuerpo


def test_reportes_preview_inventario(client, monkeypatch):
    _login(client, monkeypatch)
    _stub(monkeypatch)
    cuerpo = client.get("/reportes?tipo=inventario").get_data(as_text=True)
    assert "Reporte de Inventario" in cuerpo
    assert "Café" in cuerpo
    # inventario no lleva fila de totales (total_row=None): no debe fallar el
    # render y no debe aparecer una fila "Total" espuria.
    assert re.search(r"<td[^>]*>Total</td>", cuerpo) is None


def test_reportes_preview_estado_resultados(client, monkeypatch):
    _login(client, monkeypatch)
    _stub(monkeypatch)
    cuerpo = client.get("/reportes?tipo=estado_resultados").get_data(as_text=True)
    assert "Estado de Resultados" in cuerpo
    assert "Utilidad" in cuerpo
    assert re.search(r"<td[^>]*>Total</td>", cuerpo)
    assert "600.00" in cuerpo  # utilidad de la fila
    assert "600.00" in cuerpo  # utilidad en el total (única fila -> mismo valor)


def test_sidebar_incluye_reportes(client, monkeypatch):
    _login(client, monkeypatch)
    _stub(monkeypatch)
    cuerpo = client.get("/reportes?tipo=ventas").get_data(as_text=True)
    # El enlace del sidebar renderiza href="{{ url_for('reportes.index') }}" -> /reportes.
    # Verificamos el ancla real, no solo el texto "Reportes" (presente también en <title>/<h1>).
    assert 'href="/reportes"' in cuerpo


def test_reportes_filtros_ventas_llegan_al_api_client(client, monkeypatch):
    capturados = _stub(monkeypatch)
    _login(client, monkeypatch)
    client.get("/reportes?tipo=ventas&id_usuario=7&id_metodo=2")
    assert capturados["ventas"] == {"id_usuario": 7, "id_metodo": 2}


def test_reportes_filtros_gastos_llegan_al_api_client(client, monkeypatch):
    capturados = _stub(monkeypatch)
    _login(client, monkeypatch)
    client.get("/reportes?tipo=gastos&id_usuario=7&id_categoria=3")
    assert capturados["gastos"] == {"id_usuario": 7, "id_categoria": 3}


def test_reportes_filtro_inventario_bajo_minimo(client, monkeypatch):
    capturados = _stub(monkeypatch)
    _login(client, monkeypatch)
    client.get("/reportes?tipo=inventario&solo_bajo_minimo=1")
    assert capturados["inventario"] == {"solo_bajo_minimo": True}
    # Sin el parámetro, debe ir False (comportamiento por defecto del dashboard).
    client.get("/reportes?tipo=inventario")
    assert capturados["inventario"] == {"solo_bajo_minimo": False}


def test_reportes_filtro_estado_resultados_agrupar(client, monkeypatch):
    capturados = _stub(monkeypatch)
    _login(client, monkeypatch)
    client.get("/reportes?tipo=estado_resultados&agrupar=mes")
    assert capturados["estado_resultados"] == {"agrupar": "mes"}


def test_reportes_dropdowns_poblados(client, monkeypatch):
    _login(client, monkeypatch)
    _stub(monkeypatch)
    cuerpo = client.get("/reportes?tipo=ventas").get_data(as_text=True)
    assert "Juan Pérez" in cuerpo
    assert "Tarjeta" in cuerpo
    cuerpo_gastos = client.get("/reportes?tipo=gastos").get_data(as_text=True)
    assert "Servicios" in cuerpo_gastos


def test_reportes_contenedores_data_para_y_toggle_js(client, monkeypatch):
    _login(client, monkeypatch)
    _stub(monkeypatch)
    cuerpo = client.get("/reportes?tipo=ventas").get_data(as_text=True)
    assert 'id="tipo-reporte"' in cuerpo
    assert 'data-para="ventas gastos estado_resultados"' in cuerpo
    assert 'data-para="ventas gastos"' in cuerpo
    assert 'data-para="ventas"' in cuerpo
    assert 'data-para="gastos"' in cuerpo
    assert 'data-para="estado_resultados"' in cuerpo
    assert 'data-para="inventario"' in cuerpo
    assert "is-hidden" in cuerpo
    assert "aplicarToggleFiltros" in cuerpo
    assert "onchange=\"this.form.submit()\"" not in cuerpo


def test_reportes_estado_resultados_muestra_grafica(client, monkeypatch):
    _login(client, monkeypatch)
    _stub(monkeypatch)
    cuerpo = client.get("/reportes?tipo=estado_resultados").get_data(as_text=True)
    assert 'id="chart-er"' in cuerpo
    assert 'class="chart-box"' in cuerpo
    assert "maintainAspectRatio: false" in cuerpo
    match = re.search(r'<script id="er-data"[^>]*>(.*?)</script>', cuerpo, re.S)
    assert match is not None
    assert "2026-07" in match.group(1)
    assert "utilidad" in match.group(1)


def test_reportes_grafica_ausente_en_otros_tipos(client, monkeypatch):
    _login(client, monkeypatch)
    _stub(monkeypatch)
    for tipo in ("ventas", "gastos", "inventario"):
        cuerpo = client.get(f"/reportes?tipo={tipo}").get_data(as_text=True)
        assert 'id="chart-er"' not in cuerpo


def test_reportes_js_no_declara_globales_reservados(client, monkeypatch):
    # Mismo aprendizaje que en el dashboard: declarar `const top`/name/parent/
    # self/location/... en el <script> global rompe TODO el script en el
    # navegador (no lo detecta node --check, solo el runtime del browser).
    _login(client, monkeypatch)
    _stub(monkeypatch)
    cuerpo = client.get("/reportes?tipo=estado_resultados").get_data(as_text=True)
    reservados = (
        "top", "parent", "self", "name", "location", "length",
        "closed", "frames", "origin", "status", "window",
    )
    patron = re.compile(
        r"\b(?:const|let|var)\s+(" + "|".join(reservados) + r")\b"
    )
    encontrados = patron.findall(cuerpo)
    assert not encontrados, f"nombres globales reservados declarados: {encontrados}"


XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_reportes_export_xlsx(client, monkeypatch):
    _login(client, monkeypatch)
    _stub(monkeypatch)
    r = client.get("/reportes?tipo=ventas&preset=rango&desde=2026-07-01&hasta=2026-07-31&formato=xlsx")
    assert r.status_code == 200
    assert r.mimetype == XLSX_CT
    assert "attachment" in r.headers["Content-Disposition"]
    assert r.data[:2] == b"PK"  # los .xlsx son ZIP


def test_reportes_export_pdf(client, monkeypatch):
    _login(client, monkeypatch)
    _stub(monkeypatch)
    r = client.get("/reportes?tipo=ventas&preset=rango&desde=2026-07-01&hasta=2026-07-31&formato=pdf")
    assert r.status_code == 200
    assert r.mimetype == "application/pdf"
    assert "attachment" in r.headers["Content-Disposition"]
    assert r.data[:4] == b"%PDF"


def test_reportes_xlsx_celdas_numericas(client, monkeypatch):
    _login(client, monkeypatch)
    _stub(monkeypatch)
    from io import BytesIO
    from openpyxl import load_workbook
    r = client.get("/reportes?tipo=ventas&preset=rango&desde=2026-07-01&hasta=2026-07-31&formato=xlsx")
    ws = load_workbook(BytesIO(r.data)).active
    # data row 2, "Total" column (4th) must be a real number, not "232.00"
    assert ws.cell(row=2, column=4).value == 232.0
    assert isinstance(ws.cell(row=2, column=4).value, (int, float))


def test_reportes_export_xlsx_inventario_sin_total_row(client, monkeypatch):
    # inventario tiene total_row=None: el exportador no debe reventar al no
    # anexar fila de totales.
    _login(client, monkeypatch)
    _stub(monkeypatch)
    r = client.get("/reportes?tipo=inventario&formato=xlsx")
    assert r.status_code == 200
    assert r.mimetype == XLSX_CT
    assert r.data[:2] == b"PK"


def test_reportes_export_pdf_inventario_sin_total_row(client, monkeypatch):
    _login(client, monkeypatch)
    _stub(monkeypatch)
    r = client.get("/reportes?tipo=inventario&formato=pdf")
    assert r.status_code == 200
    assert r.mimetype == "application/pdf"
    assert r.data[:4] == b"%PDF"
